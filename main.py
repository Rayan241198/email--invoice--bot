# --- Standard imports ---
import os
import re
import imaplib
from imaplib import IMAP4_SSL
import email
from email.header import decode_header, make_header
from email.policy import default
from email.utils import parseaddr
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
import getpass

# --- Optional ML import (kept; guarded by try/except when called) ---
from ai_fraud_ml import predict_email_risk

# =========================
# Helpers (ML + parsing)
# =========================
def domain_from_header(header_value: str) -> str:
    """'Name <user@acme.com>' -> 'acme.com'."""
    name, email_addr = parseaddr(header_value or "")
    return email_addr.split("@")[1].lower() if email_addr and "@" in email_addr else ""

def attachment_types_from_list(names):
    """Return comma-separated lowercase file extensions from attachment names."""
    exts = []
    for n in names or []:
        _, ext = os.path.splitext((n or "").strip())
        if ext:
            exts.append(ext.lstrip(".").lower())
    return ",".join(sorted(set(exts))) if exts else ""

def extract_amount_guess(text: str):
    """Very simple amount extractor: looks for a number like 123.45 in the email body."""
    if not text:
        return None
    m = re.search(r'([0-9]+(?:\.[0-9]{2})?)', text)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None

def decode_mime_header(raw_value):
    if not raw_value:
        return ""
    try:
        return str(make_header(decode_header(raw_value)))
    except Exception:
        return str(raw_value)

def extract_text_body(msg):
    if msg.is_multipart():
        parts = []
        for part in msg.walk():
            ctype = part.get_content_type()
            disp = str(part.get("Content-Disposition", ""))
            if ctype == "text/plain" and "attachment" not in disp.lower():
                try:
                    parts.append(part.get_content())
                except Exception:
                    payload = part.get_payload(decode=True) or b""
                    parts.append(payload.decode(errors="ignore"))
        return "\n".join(parts).strip()
    else:
        if msg.get_content_type() == "text/plain":
            try:
                return msg.get_content()
            except Exception:
                return (msg.get_payload(decode=True) or b"").decode(errors="ignore")
        return ""

def list_attachments(msg, save=False, save_dir=Path("attachments")):
    out = []
    for part in msg.walk():
        disp = str(part.get("Content-Disposition", ""))
        if "attachment" in disp.lower():
            raw_name = part.get_filename()
            fname = decode_mime_header(raw_name) or "unnamed"
            data = part.get_payload(decode=True)
            size = len(data) if data else 0
            ctype = part.get_content_type()
            is_pdf = (ctype == "application/pdf") or fname.lower().endswith(".pdf")
            if save and data:
                try:
                    save_dir.mkdir(exist_ok=True)
                    (save_dir / fname).write_bytes(data)
                except Exception as e:
                    print("Couldn't save attachment", fname, e)
            out.append((fname, size, is_pdf, data if save else None))
    return out

# =========================
# Simple invoice heuristic
# =========================
KEYWORDS = {"invoice", "payment", "receipt", "bill", "statement", "due", "amount due", "balance", "order", "paid", "unpaid"}
INVOICE_RE = re.compile(r"(?i)(invoice\s*#?\s*\d{3,}|inv[-\s]?\d{3,}|#\d{3,})")

def looks_like_invoice(subject, body, attachments):
    text = f"{subject}\n{body}".lower()
    kw_hit = "invoice" in text  # minimal heuristic
    is_invoice = kw_hit
    reason = "keyword" if kw_hit else "no strong signal"
    return is_invoice, reason

# =========================
# Excel helpers
# =========================
def ensure_workbook(path: Path):
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    ws.append([
        "SavedAt","From","Subject","Date","HasPDF","AttachmentNames","Reason","MessageID",
        "FromDomain","ReplyDomain","AttachmentTypes","AmountGuess",
        "ML Risk Score","ML Top Tokens"
    ])
    wb.save(path)
    return wb

def append_invoice_row(path: Path, *, from_, subject, date_str, has_pdf, attach_names, reason, message_id,
                       from_domain="", reply_domain="", attachment_types="", amount_guess="",
                       ml_risk_score="", ml_top_tokens=""):
    wb = ensure_workbook(path); ws = wb["Invoices"]
    saved_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws.append([
        saved_at, from_, subject, date_str, has_pdf, ", ".join(attach_names), reason, message_id,
        from_domain, reply_domain, attachment_types, amount_guess,
        ml_risk_score, ml_top_tokens
    ])
    wb.save(path)

# =========================
# IMAP keep-alive
# =========================
def maybe_keepalive(imap, i, every=50):
    """Ping the server every N messages to avoid disconnects."""
    if i % every == 0:
        try:
            imap.noop()
        except imaplib.IMAP4.abort:
            # Optionally implement reconnect here
            pass

# =========================
# Settings / Inputs
# =========================
print("=== Email Invoice Bot ===")
GMAIL_ADDRESS = input("Enter your Gmail address: ").strip()
APP_PASSWORD = getpass.getpass("Enter your 16-char Gmail App Password: ").strip()

IMAP_SERVER = "imap.gmail.com"
MAILBOX = "INBOX"
EXCEL_PATH = Path("invoices.xlsx")
SAVE_ATTACHMENTS = False
ATTACH_DIR = Path("attachments")

# =========================
# Main
# =========================
def process_all(limit: int = 200):
    with IMAP4_SSL(IMAP_SERVER) as imap:
        imap.login(GMAIL_ADDRESS, APP_PASSWORD)
        imap.select(MAILBOX, readonly=True)

        SEARCH_CRITERIA = 'ALL'
        status, data = imap.search(None, SEARCH_CRITERIA)
        if status != "OK":
            print("Search failed:", status)
            return

        msg_ids = data[0].split()
        print(f"Found {len(msg_ids)} messages matching {SEARCH_CRITERIA}.")
        # Scan the most recent N
        msg_ids = msg_ids[-min(limit, 200):]

        count = 0
        for i, msg_id in enumerate(reversed(msg_ids), 1):
            maybe_keepalive(imap, i, every=50)

            status, msg_data = imap.fetch(msg_id, "(RFC822)")
            if status != "OK":
                print("Fetch failed for", msg_id)
                continue

            raw_bytes = msg_data[0][1]
            msg = email.message_from_bytes(raw_bytes, policy=default)

            subject = decode_mime_header(msg.get("Subject"))
            from_   = decode_mime_header(msg.get("From"))
            date_h  = decode_mime_header(msg.get("Date"))
            body = extract_text_body(msg)
            attachments = list_attachments(msg, save=SAVE_ATTACHMENTS, save_dir=ATTACH_DIR)

            is_inv, reason = looks_like_invoice(subject, body, attachments)

            has_pdf = any(is_pdf for _, _, is_pdf, _ in attachments)
            attach_names = [a[0] for a in attachments] if attachments else []
            message_id = decode_mime_header(msg.get("Message-ID"))

            print("DEBUG:", subject, "| From:", from_)

            if is_inv:
                # Build ML inputs
                from_domain    = domain_from_header(from_)
                reply_domain   = domain_from_header(decode_mime_header(msg.get("Reply-To")))
                attachment_types = attachment_types_from_list(attach_names)
                amount_value   = extract_amount_guess(f"{subject}\n{body}")

                # Call ML safely (in case the module/return shape changes)
                try:
                    ml = predict_email_risk(
                        subject=subject,
                        body=body,
                        from_domain=from_domain,
                        reply_domain=reply_domain,
                        attachment_types=attachment_types,
                        amount=amount_value
                    )
                    ml_score  = ml.get("risk_score", "")
                    ml_tokens = ", ".join(ml.get("top_tokens", []))
                except Exception:
                    ml_score, ml_tokens = "", ""

                append_invoice_row(
                    EXCEL_PATH,
                    from_=from_,
                    subject=subject,
                    date_str=date_h,
                    has_pdf=has_pdf,
                    attach_names=attach_names,
                    reason=reason,
                    message_id=message_id,
                    from_domain=from_domain,
                    reply_domain=reply_domain,
                    attachment_types=attachment_types,
                    amount_guess=amount_value if amount_value is not None else "",
                    ml_risk_score=ml_score,
                    ml_top_tokens=ml_tokens
                )

                print(f"✔ Saved invoice: {subject!r} | reason: {reason} | ML Risk={ml_score}")
                count += 1
            else:
                print(f"· Skipped: {subject!r}")

        # ONE safe logout (with `with` it auto-closes anyway)
        try:
            imap.logout()
        except Exception:
            pass

        if count == 0:
            print("No invoices matched. Saved invoices.xlsx")

if __name__ == "__main__":
    print(">>> RUNNING NEW BOT VERSION <<<")
    process_all(limit=200)
